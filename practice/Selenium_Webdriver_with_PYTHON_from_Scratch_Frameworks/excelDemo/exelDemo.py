import openpyxl

mybook = openpyxl.load_workbook("E:\\11)_Eleusis_Git_Stuf\\WebAutomation\\PyhtonDemo.xlsx")
sheet =  mybook.active
# get specifically the rows and olum coordinate of a cell
cell = sheet.cell(row=1, column=1)
print(f"My cell: {cell.value}")

# Set some info to the file
# sheet.cell(row=2, column=2).value = "Eleusis"
print(sheet.cell(row=2, column=2).value)

#Get maximun row and colums

print(f"MAX Row {sheet.max_row}")
print(f"MAX Column {sheet.max_column}")

# Get info by cells coordinates

print(f"My cell: {sheet['A5'].value}")

# Get values from the whole file
print("____________________WHOLE_SHEET_____________________")

# for i in range(1, sheet.max_row + 1):
#     for j in range(1, sheet.max_column +1):
#         print(sheet.cell(row=i, column=j).value)
       

tmp_excel_dictionary= {}
whole_excle_data = []

tmp_keys = []
for tmp_row in range(1, sheet.max_row + 1):
    for tmp_column in range(1, sheet.max_column +1):
        if tmp_row == 1:
            tmp_keys.append(sheet.cell(row=tmp_row, column=tmp_column).value)
        else:
            tmp_excel_dictionary.update({tmp_keys[tmp_column - 1]:sheet.cell(row=tmp_row, column=tmp_column).value})
    if tmp_excel_dictionary:
        whole_excle_data.append(tmp_excel_dictionary.copy())

print(f"Final Excel info in diccionary: {whole_excle_data}")