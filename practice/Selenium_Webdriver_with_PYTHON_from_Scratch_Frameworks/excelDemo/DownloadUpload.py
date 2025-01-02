from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
import openpyxl


class ExcelFileManagerError(Exception):
    pass


class ExcelFileManager:
    def __init__(self, excel_path):
        self._excel_path = excel_path
        self.excel_book =  self.load_file()

    @property
    def excel_path(self) -> str:
        return self._excel_path

    @excel_path.setter
    def excel_path(self, new_path:str) -> None:
        if not isinstance(new_path, str):
            raise ExcelFileManagerError("Wrong type for excel path, it should a str")
        self._excel_path = new_path
    
    def load_file(self):
        return openpyxl.load_workbook(self.excel_path)
    
    def edit_cell_book(self, column_filed, row_filed, new_value):
        sheet =  self.excel_book.active
        def find_coordinate():
            x_coo = None
            y_coo = None
            for row in range(1, sheet.max_row + 1):
                for column in range(1, sheet.max_column):
                    if sheet.cell(row=1, column=column).value == column_filed:
                        x_coo = column
                    if sheet.cell(row=row, column=column).value == row_filed:
                        y_coo = row
                    if x_coo and y_coo:
                        break
                if x_coo and y_coo:
                    break
            return y_coo, x_coo
        sheet.cell(*find_coordinate()).value = new_value
        self.excel_book.save(self.excel_path)

if __name__ == '__main__':
    file_path = "C:\\Users\\eleus\\Downloads\\download.xlsx"
    driver = webdriver.Chrome()
    driver.implicitly_wait(5)
    driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
    driver.find_element(By.ID, "downloadButton").click()
    new_value = "90"
    fruit_name = "Apple"
    local_exc_manager = ExcelFileManager(file_path)
    local_exc_manager.edit_cell_book(column_filed="price",row_filed=fruit_name, new_value=new_value)
    
    file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
    file_input.send_keys(file_path)

    toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
    wait = WebDriverWait(driver, 5)
    succesfully_msg = wait.until(expected_conditions.visibility_of_element_located(toast_locator)).text
    print(succesfully_msg)
    priceColumn = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
    actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
    assert actual_price == new_value
